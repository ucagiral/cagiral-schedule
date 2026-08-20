#!/usr/bin/env python3
"""Offline checks for everything in tools/loops that does not need the network.

Run by CI before any query goes out, so a parsing or rendering regression is
caught without waiting on a portal.  The layers that genuinely require remote
data (contact matrices) are exercised through their pure helpers only.

    python3 tools/loops/selftest.py
"""

from __future__ import annotations

import os
import sys
import tempfile
import traceback
from typing import Any, Callable

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import layers  # noqa: E402
import query  # noqa: E402
import sources  # noqa: E402
import workbook as workbook_module  # noqa: E402
from layers import Region  # noqa: E402

FAILURES: list[str] = []


def check(name: str, condition: bool, detail: str = "") -> None:
    if condition:
        print(f"  ok   {name}")
    else:
        print(f"  FAIL {name} {detail}")
        FAILURES.append(f"{name} {detail}".strip())


def case(fn: Callable[[], None]) -> Callable[[], None]:
    def wrapped() -> None:
        print(f"\n{fn.__name__}")
        try:
            fn()
        except Exception:  # noqa: BLE001 - a raising case is a failing case
            print(traceback.format_exc())
            FAILURES.append(f"{fn.__name__} raised")
    return wrapped


# --------------------------------------------------------------------------

@case
def test_span_and_locus_parsing() -> None:
    check("500kb", query.parse_span("500kb") == 500_000)
    check("1.5mb", query.parse_span("1.5mb") == 1_500_000)
    check("bare bp", query.parse_span("250000") == 250_000)
    check("comma bp", query.parse_span("250,000") == 250_000)

    region, origin = query.parse_locus("chrX:67,543,900-67,548,900", "hg38")
    check("coords chrom", region.chrom == "chrX", region.chrom)
    check("coords start", region.start == 67_543_900)
    check("coords end", region.end == 67_548_900)
    check("coords origin", origin == "coordinates")

    bare, _ = query.parse_locus("X:100-200", "hg38")
    check("chr prefix added", bare.chrom == "chrX", bare.chrom)

    for bad in ("chrX:200-100", "not a locus at all!!"):
        try:
            query.parse_locus(bad, "hg38")
        except (ValueError, sources.SourceError):
            check(f"rejects {bad!r}", True)
        else:
            check(f"rejects {bad!r}", False, "accepted it")


@case
def test_chrom_normalisation_and_overlap() -> None:
    check("chr prefix ignored", layers.normalise_chrom("chrX") == layers.normalise_chrom("X"))
    check("MT folds to M", layers.normalise_chrom("MT") == "M")
    region = Region("chrX", 1_000, 2_000)
    check("overlap inside", region.overlaps("X", 1_500, 1_600))
    check("overlap touching left", region.overlaps("chrX", 900, 1_001))
    check("no overlap abutting", not region.overlaps("chrX", 2_000, 2_100))
    check("slop reaches", region.overlaps("chrX", 2_500, 2_600, slop=1_000))
    check("other chromosome", not region.overlaps("chr7", 1_500, 1_600))


@case
def test_loop_scanning() -> None:
    bedpe = "\n".join([
        "#chr1\tx1\tx2\tchr2\ty1\ty2\tobserved\tfdr_donut",
        # anchor 1 on the query, anchor 2 on the partner -- the wanted hit
        "chrX\t66895000\t66900000\tchrX\t67540000\t67545000\t41\t3.1e-4",
        # touches the query but lands somewhere else entirely
        "chrX\t67540000\t67545000\tchrX\t70000000\t70005000\t12\t0.02",
        # nowhere near either locus
        "chrX\t10000000\t10005000\tchrX\t11000000\t11005000\t8\t0.5",
        # a different chromosome
        "chr7\t67540000\t67545000\tchr7\t66895000\t66900000\t99\t1e-9",
    ])
    original = layers.fetch_lines
    layers.fetch_lines = lambda url: iter(bedpe.split("\n"))  # type: ignore[assignment]
    try:
        pf = sources.PortalFile(accession="TEST01", portal="TEST", url="memory://",
                                file_format="bedpe", file_type="loops",
                                genome="hg38", cell_type="LNCaP", assay="in situ Hi-C")
        q = Region("chrX", 67_543_900, 67_548_900)
        p = Region("chrX", 66_895_000, 66_904_000)

        paired = layers.scan_loop_file(pf, q, p, slop=0)
        check("one paired loop", len(paired) == 1, f"got {len(paired)}")
        if paired:
            row = paired[0]
            check("fdr read from header", row["fdr"] == 3.1e-4, str(row["fdr"]))
            check("score read from header", row["score"] == 41.0, str(row["score"]))
            check("separation", row["separation_bp"] == 645_000, str(row["separation_bp"]))
            check("resolution", row["resolution_bp"] == 5_000)

        unpaired = layers.scan_loop_file(pf, q, None, slop=0)
        check("two loops touch the query", len(unpaired) == 2, f"got {len(unpaired)}")
    finally:
        layers.fetch_lines = original  # type: ignore[assignment]


@case
def test_loop_scanning_without_header() -> None:
    bedpe = "chrX\t66895000\t66900000\tchrX\t67540000\t67545000"
    original = layers.fetch_lines
    layers.fetch_lines = lambda url: iter(bedpe.split("\n"))  # type: ignore[assignment]
    try:
        pf = sources.PortalFile(accession="TEST02", portal="TEST", url="memory://",
                                file_format="bedpe", file_type="loops", genome="hg38")
        rows = layers.scan_loop_file(pf, Region("chrX", 67_543_900, 67_548_900),
                                     None, slop=0)
        check("headerless file still parses", len(rows) == 1, f"got {len(rows)}")
        if rows:
            check("no score invented", rows[0]["score"] is None)
            check("no fdr invented", rows[0]["fdr"] is None)
    finally:
        layers.fetch_lines = original  # type: ignore[assignment]


@case
def test_motif_scanning() -> None:
    # A toy matrix that only ever matches ACGT, so the expected hit is knowable.
    pfm = [[100.0, 0, 0, 0], [0, 100.0, 0, 0], [0, 0, 100.0, 0], [0, 0, 0, 100.0]]
    pwm = layers._pwm_from_pfm(pfm)

    score, offset, strand, relative = layers.best_motif("TTTACGTTTT", pwm)
    check("forward strand found", strand == "+", strand)
    check("forward offset", offset == 3, str(offset))
    check("relative score high", relative > 0.95, f"{relative:.3f}")

    # revcomp(ACGT) is ACGT, so use a palindrome-free probe for the minus strand.
    pfm2 = [[100.0, 0, 0, 0], [100.0, 0, 0, 0], [0, 0, 100.0, 0], [0, 0, 0, 100.0]]
    pwm2 = layers._pwm_from_pfm(pfm2)          # matches AAGT, revcomp ACTT
    _, offset2, strand2, _ = layers.best_motif("TTTACTTGGG", pwm2)
    check("reverse strand found", strand2 == "-", strand2)
    check("reverse offset in forward coords", offset2 == 3, str(offset2))

    _, offset3, _, _ = layers.best_motif("NNNN", pwm)
    check("all-N sequence yields no hit", offset3 == -1, str(offset3))


@case
def test_orientation_is_judged_in_genomic_order() -> None:
    def row(anchor: str, start: int, strand: str) -> dict[str, Any]:
        return {"anchor": anchor, "peak_start": start, "strand": strand,
                "motif_rel_score": 0.9}

    # "partner" sorts before "query" alphabetically but sits downstream, so a
    # label-ordered implementation would call this convergent pair divergent.
    convergent = [row("query", 1_000, "+"), row("partner", 900_000, "-")]
    check("convergent", layers.orientation_verdict(convergent) == "convergent",
          layers.orientation_verdict(convergent))

    divergent = [row("query", 1_000, "-"), row("partner", 900_000, "+")]
    check("divergent", layers.orientation_verdict(divergent) == "divergent",
          layers.orientation_verdict(divergent))

    tandem = [row("query", 1_000, "+"), row("partner", 900_000, "+")]
    check("tandem", layers.orientation_verdict(tandem) == "tandem",
          layers.orientation_verdict(tandem))

    check("unknown with one anchor",
          layers.orientation_verdict([row("query", 1_000, "+")]) == "unknown")
    check("unknown when strand missing",
          layers.orientation_verdict([row("query", 1, "?"), row("partner", 2, "?")])
          == "unknown")


@case
def test_pair_enrichment_picks_the_right_bins() -> None:
    q = Region("chrX", 67_543_900, 67_548_900)   # mid 67,546,400 -> bin 67,540,000
    p = Region("chrX", 66_895_000, 66_904_000)   # mid 66,899,500 -> bin 66,895,000
    rows = [
        {"dataset": "D1", "cell_type": "LNCaP", "normalisation": "KR",
         "resolution_bp": 5_000, "chrom": "chrX",
         "bin1_start": 66_895_000, "bin2_start": 67_540_000,
         "observed": 41.0, "expected": 8.2, "oe": 5.0, "source_url": "u"},
        {"dataset": "D1", "cell_type": "LNCaP", "normalisation": "KR",
         "resolution_bp": 5_000, "chrom": "chrX",
         "bin1_start": 66_800_000, "bin2_start": 67_100_000,
         "observed": 3.0, "expected": 3.1, "oe": 0.97, "source_url": "u"},
    ]
    hits = layers.pair_enrichment(rows, q, p)
    check("one entry per dataset", len(hits) == 1, f"got {len(hits)}")
    if hits:
        check("carries the O/E", hits[0]["oe"] == 5.0)
        check("counted one bin pair", hits[0]["bin_pairs"] == 1, str(hits[0]["bin_pairs"]))

    # A locus straddling a bin edge must still match: the query spans bins
    # 67,540,000 and 67,545,000, so both pair with the partner bin.
    spanning = rows[:1] + [dict(rows[0], bin2_start=67_545_000, oe=6.5, observed=55.0)]
    hits2 = layers.pair_enrichment(spanning, q, p)
    check("straddling bins collapse to one row", len(hits2) == 1, f"got {len(hits2)}")
    if hits2:
        check("strongest bin pair wins", hits2[0]["oe"] == 6.5, str(hits2[0]["oe"]))
        check("both bin pairs counted", hits2[0]["bin_pairs"] == 2,
              str(hits2[0]["bin_pairs"]))


@case
def test_workbook_roundtrip() -> None:
    loop_rows = [{
        "cell_type": "LNCaP", "assay": "in situ Hi-C", "caller": "loops",
        "dataset": "TEST01", "portal": "TEST", "genome": "hg38",
        "chrom1": "chrX", "start1": 66_895_000, "end1": 66_900_000,
        "chrom2": "chrX", "start2": 67_540_000, "end2": 67_545_000,
        "separation_bp": 648_000, "resolution_bp": 5_000, "score": 41.0,
        "fdr": 3.1e-4, "anchor1_hits_query": False, "anchor2_hits_query": True,
        "source_url": "memory://",
    }]
    contact_rows = [{
        "cell_type": "LNCaP", "dataset": "TEST01", "portal": "TEST",
        "normalisation": "KR", "resolution_bp": 5_000, "chrom": "chrX",
        "bin1_start": 66_895_000, "bin2_start": 67_540_000, "observed": 41.0,
        "expected": 8.2, "oe": 5.0, "source_url": "memory://",
    }]

    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "roundtrip.xlsx")
        counts = workbook_module.write_workbook(
            path, loops=loop_rows, contact_matrix=contact_rows, tads=[],
            ctcf=[], datasets=[])
        check("loops counted", counts["loops"] == 1, str(counts))
        check("contact counted", counts["contact_matrix"] == 1, str(counts))
        check("empty sheets are zero", counts["tads"] == 0 and counts["ctcf"] == 0)

        problems = workbook_module.verify_workbook(path, counts)
        check("verification clean", not problems, str(problems))

        # A wrong expectation must be caught, or the CI check is decorative.
        bad = dict(counts, loops=99)
        check("verification catches a miscount",
              bool(workbook_module.verify_workbook(path, bad)))

        from openpyxl import load_workbook
        book = load_workbook(path, read_only=True)
        try:
            check("five sheets", len(book.sheetnames) == 5, str(book.sheetnames))
            check("sheet order stable",
                  book.sheetnames == list(workbook_module.SHEETS),
                  str(book.sheetnames))
            values = [c.value for c in next(book["loops"].iter_rows(min_row=2, max_row=2))]
            check("fdr survives as a number", values[15] == 3.1e-4, str(values[15]))
        finally:
            book.close()


@case
def test_report_and_workbook_agree() -> None:
    loop_rows = [{
        "cell_type": f"CELL{i}", "assay": "Hi-C", "caller": "loops",
        "dataset": f"D{i}", "portal": "TEST", "genome": "hg38",
        "chrom1": "chrX", "start1": 1_000 * i, "end1": 1_000 * i + 500,
        "chrom2": "chrX", "start2": 900_000, "end2": 905_000,
        "separation_bp": 900_000, "resolution_bp": 5_000, "score": 1.0,
        "fdr": 0.001, "anchor1_hits_query": True, "anchor2_hits_query": False,
        "source_url": "memory://",
    } for i in range(1, 31)]

    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "agree.xlsx")
        counts = workbook_module.write_workbook(
            path, loops=loop_rows, contact_matrix=[], tads=[], ctcf=[], datasets=[])
        check("all 30 loops in the workbook", counts["loops"] == 30, str(counts))

        report = query.render_markdown(
            title="t", query=Region("chrX", 1_000, 2_000), partner=None,
            origin="coordinates", build="hg38", catalogue_size=1, cell_count=30,
            verdict=["v"], loop_rows=loop_rows, enrichment=[], tad_rows=[],
            ctcf_rows=[], notes=[], counts=counts, workbook_name="agree.xlsx",
            generated="now")
        # The report shows the first 25 and says so; the workbook holds them all.
        check("report truncation is stated", "5 further loops" in report)
        check("workbook is linked", "agree.xlsx" in report)
        check("amplification caveat present", "Copy number inflates" in report)
        check("no-call caveat present", "Absence of a call" in report)


@case
def test_verdict_wording() -> None:
    q, p = Region("chrX", 67_543_900, 67_548_900), Region("chrX", 66_895_000, 66_904_000)

    loops = [{"cell_type": "LNCaP", "separation_bp": 648_000}]
    said = " ".join(query.build_verdict(loops, [], [], [], q, p))
    check("called loop reported", "Yes in LNCaP" in said, said)

    enriched = [{"cell_type": "LNCaP", "dataset": "D1", "oe": 4.2}]
    said = " ".join(query.build_verdict([], enriched, [], [], q, p))
    check("elevated contact reported", "4.2×" in said, said)
    check("and marked as uncalled", "No caller flagged" in said, said)

    weak = [{"cell_type": "LNCaP", "dataset": "D1", "oe": 1.05}]
    said = " ".join(query.build_verdict([], weak, [], [], q, p))
    check("weak contact is not a loop", "No called loop and no elevated" in said, said)

    boundary = [{"cell_type": "LNCaP", "between_query_and_partner": True,
                 "contains_query": False, "contains_partner": False}]
    said = " ".join(query.build_verdict([], [], boundary, [], q, p))
    check("boundary raises suspicion", "suspicion" in said, said)

    said = " ".join(query.build_verdict([], [], [], [], q, None))
    check("no-partner wording", "No called loop has an anchor" in said, said)


@case
def test_portal_fields_survive_inconsistent_arity() -> None:
    # ENCODE returns assay_term_name as a string on some files and a list on
    # others; the list form used to crash classification with
    # "sequence item 2: expected str instance, list found".
    pf = sources.PortalFile(
        accession="ENCFF000AAA", portal="ENCODE", url="u", file_format="bedpe",
        file_type="loops", genome="hg38", cell_type="LNCaP",
        assay=["Hi-C", "in situ Hi-C"], description=None)  # type: ignore[arg-type]
    check("assay list flattened", pf.assay == "Hi-C in situ Hi-C", pf.assay)
    check("None description becomes empty", pf.description == "")
    check("haystack builds", "hi-c" in pf.haystack, pf.haystack)
    check("classification still works", sources.is_loop_file(pf))

    sized = sources.PortalFile(accession="A", portal="P", url="u",
                               file_format="hic", file_type="contact matrix",
                               genome="hg38", size=None)  # type: ignore[arg-type]
    check("None size becomes zero", sized.size == 0)

    nested = sources.PortalFile(accession="A", portal="P", url="u",
                                file_format="bed", file_type=["IDR peaks", None],
                                genome="hg38")  # type: ignore[arg-type]
    check("None inside a list is dropped", nested.file_type == "IDR peaks",
          nested.file_type)


@case
def test_file_classification() -> None:
    def pf(**kw: Any) -> sources.PortalFile:
        base = dict(accession="A", portal="P", url="u", file_format="bedpe",
                    file_type="", genome="hg38")
        base.update(kw)
        return sources.PortalFile(**base)  # type: ignore[arg-type]

    check("hiccups bedpe is a loop file",
          sources.is_loop_file(pf(file_type="loops")))
    check("chromatin interactions is a loop file",
          sources.is_loop_file(pf(file_type="chromatin interactions")))
    check("a bare bed is not a loop file",
          not sources.is_loop_file(pf(file_format="bed", file_type="peaks")))
    check("hic is a matrix", sources.is_contact_matrix(pf(file_format="hic")))
    check("mcool is a matrix", sources.is_contact_matrix(pf(file_format="mcool")))
    check("boundaries bed", sources.is_boundary_file(
        pf(file_format="bed", file_type="boundaries")))
    check("insulation bed", sources.is_boundary_file(
        pf(file_format="bed", file_type="insulation score")))
    check("ctcf idr peaks", sources.is_ctcf_peaks(
        pf(file_format="bed narrowPeak", file_type="IDR thresholded peaks",
           assay="ChIP-seq", cell_type="LNCaP", description="CTCF")))
    check("h3k27ac peaks are not ctcf", not sources.is_ctcf_peaks(
        pf(file_format="bed narrowPeak", file_type="IDR thresholded peaks",
           assay="ChIP-seq", cell_type="LNCaP", description="H3K27ac")))


@case
def test_cap_is_spread_across_cell_types() -> None:
    def pf(cell: str, n: int) -> sources.PortalFile:
        return sources.PortalFile(accession=f"{cell}{n}", portal="P", url="u",
                                  file_format="bedpe", file_type="loops",
                                  genome="hg38", cell_type=cell)

    # A549 sorts first and LNCaP late; a prefix truncation would keep only A549.
    files = [pf("A549", n) for n in range(20)] + [pf("LNCaP", n) for n in range(20)]
    picked, capped = layers._select(files, [], cap=4)
    check("cap reported", capped)
    check("cap respected", len(picked) == 4, str(len(picked)))
    cells = {p.cell_type for p in picked}
    check("both cell types survive the cap", cells == {"A549", "LNCaP"}, str(cells))

    # Under the cap nothing is dropped or reordered away.
    picked, capped = layers._select(files, [], cap=100)
    check("no cap when it fits", not capped and len(picked) == 40, str(len(picked)))

    # An explicit filter still narrows to the named line.
    picked, _ = layers._select(files, ["LNCaP"], cap=100)
    check("filter narrows", {p.cell_type for p in picked} == {"LNCaP"})

    # A lopsided pool must not starve the smaller cell type.
    lopsided = [pf("A549", n) for n in range(50)] + [pf("LNCaP", 0)]
    picked, _ = layers._select(lopsided, [], cap=5)
    check("rare cell type is not starved",
          "LNCaP" in {p.cell_type for p in picked},
          str({p.cell_type for p in picked}))


@case
def test_slug_and_cells() -> None:
    check("slug of a pair",
          query.slugify("AR", "chrX:66,895,000-66,904,000")
          == "ar-vs-chrx-66-895-000-66-904-000",
          query.slugify("AR", "chrX:66,895,000-66,904,000"))
    check("slug of one locus", query.slugify("AR", None) == "ar")
    check("slug is bounded", len(query.slugify("x" * 500, "y" * 500)) <= 80)
    check("all means no filter", query.parse_cells("all") == [])
    check("blank means no filter", query.parse_cells("") == [])
    check("list is split", query.parse_cells("LNCaP, GM12878") == ["LNCaP", "GM12878"])


# --------------------------------------------------------------------------

def main() -> int:
    for name, value in sorted(globals().items()):
        if name.startswith("test_") and callable(value):
            value()
    print()
    if FAILURES:
        print(f"{len(FAILURES)} check(s) failed:")
        for failure in FAILURES:
            print(f"  - {failure}")
        return 1
    print("all offline checks passed")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
