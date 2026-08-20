"""Render the layer output as an Excel workbook.

Five flat sheets, one row per observation.  No merged cells, no styling beyond
a frozen header, no charts: the point of this file is that Umut builds his own
pivots and plots on top of it, and anything decorative in the way of that is a
liability rather than a feature.
"""

from __future__ import annotations

from typing import Any, Sequence

# Excel's own hard ceiling; our own budget is set far below it in query.py.
EXCEL_MAX_ROWS = 1_048_576

# Column order is fixed per sheet so a workbook written today lines up with one
# written next month -- formulas and pivots built against it keep working.
SHEETS: dict[str, list[str]] = {
    "loops": [
        "cell_type", "assay", "caller", "dataset", "portal", "genome",
        "chrom1", "start1", "end1", "chrom2", "start2", "end2",
        "separation_bp", "resolution_bp", "score", "fdr",
        "anchor1_hits_query", "anchor2_hits_query", "source_url",
    ],
    "contact_matrix": [
        "cell_type", "dataset", "portal", "normalisation", "resolution_bp",
        "chrom", "bin1_start", "bin2_start", "observed", "expected", "oe",
        "source_url",
    ],
    "tads": [
        "cell_type", "dataset", "portal", "kind", "chrom", "start", "end",
        "span_bp", "boundary_strength", "contains_query", "contains_partner",
        "between_query_and_partner", "source_url",
    ],
    "ctcf": [
        "anchor", "cell_type", "dataset", "portal", "peak_chrom", "peak_start",
        "peak_end", "signal", "motif", "motif_start", "motif_end", "strand",
        "motif_score", "motif_rel_score", "source_url",
    ],
    "datasets": [
        "dataset", "portal", "cell_type", "assay", "genome", "file_type",
        "file_format", "url", "accessed_utc",
    ],
}


def _write_sheet(worksheet, columns: Sequence[str], rows: Sequence[dict[str, Any]]) -> int:
    worksheet.append(list(columns))
    written = 0
    for row in rows:
        if written + 2 > EXCEL_MAX_ROWS:
            break
        worksheet.append([_cell(row.get(name)) for name in columns])
        written += 1
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        f"A1:{_column_letter(len(columns))}{written + 1}")
    for index, name in enumerate(columns, start=1):
        worksheet.column_dimensions[_column_letter(index)].width = \
            min(42, max(11, len(name) + 3))
    return written


def _column_letter(index: int) -> str:
    from openpyxl.utils import get_column_letter
    return get_column_letter(index)


def _cell(value: Any) -> Any:
    """Coerce to something Excel stores natively, without losing precision."""
    if value is None or isinstance(value, (int, float, bool, str)):
        return value
    return str(value)


def write_workbook(path: str, *, loops: Sequence[dict[str, Any]],
                   contact_matrix: Sequence[dict[str, Any]],
                   tads: Sequence[dict[str, Any]],
                   ctcf: Sequence[dict[str, Any]],
                   datasets: Sequence[dict[str, Any]]) -> dict[str, int]:
    """Write the five sheets and return how many data rows each received."""
    from openpyxl import Workbook

    data = {
        "loops": loops,
        "contact_matrix": contact_matrix,
        "tads": tads,
        "ctcf": ctcf,
        "datasets": datasets,
    }

    workbook = Workbook()
    workbook.remove(workbook.active)  # drop the default sheet openpyxl creates
    counts: dict[str, int] = {}
    for name, columns in SHEETS.items():
        worksheet = workbook.create_sheet(title=name)
        counts[name] = _write_sheet(worksheet, columns, data[name])
    workbook.save(path)
    return counts


def verify_workbook(path: str, expected: dict[str, int]) -> list[str]:
    """Re-open a written workbook and check it says what we meant it to say.

    Run against the file on disk -- and again after a clean `git checkout` -- so
    that a `.gitattributes` slip corrupting the zip container is caught here
    rather than when somebody downloads it.
    """
    from openpyxl import load_workbook

    problems: list[str] = []
    workbook = load_workbook(path, read_only=True)
    try:
        for name, columns in SHEETS.items():
            if name not in workbook.sheetnames:
                problems.append(f"sheet {name!r} missing")
                continue
            worksheet = workbook[name]
            header = [cell.value for cell in next(worksheet.iter_rows(max_row=1))]
            if header != list(columns):
                problems.append(f"sheet {name!r} header is {header!r}")
            # max_row counts the header, so data rows are one fewer.
            actual = max(0, (worksheet.max_row or 1) - 1)
            if actual != expected.get(name, 0):
                problems.append(
                    f"sheet {name!r} holds {actual} data rows, "
                    f"expected {expected.get(name, 0)}")
    finally:
        workbook.close()
    return problems


def _main(argv: Sequence[str] | None = None) -> int:
    """`python3 tools/loops/workbook.py --check FILE` — structural check only.

    Used after committing, on a copy restored straight from git, to prove the
    stored bytes still open as a workbook.  Row counts are reported rather than
    asserted here; query.py already checked them against the report.
    """
    import argparse

    parser = argparse.ArgumentParser(description="Check a generated workbook.")
    parser.add_argument("--check", required=True, metavar="FILE")
    args = parser.parse_args(argv)

    from openpyxl import load_workbook

    book = load_workbook(args.check, read_only=True)
    try:
        problems = []
        for name, columns in SHEETS.items():
            if name not in book.sheetnames:
                problems.append(f"sheet {name!r} missing")
                continue
            worksheet = book[name]
            header = [cell.value for cell in next(worksheet.iter_rows(max_row=1))]
            if header != list(columns):
                problems.append(f"sheet {name!r} header is {header!r}")
            print(f"  {name}: {max(0, (worksheet.max_row or 1) - 1):,} data rows")
    finally:
        book.close()

    if problems:
        for problem in problems:
            print(f"  problem: {problem}")
        return 1
    print(f"{args.check} opens cleanly with all {len(SHEETS)} sheets intact")
    return 0


if __name__ == "__main__":
    raise SystemExit(_main())
